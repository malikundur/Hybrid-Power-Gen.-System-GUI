import sys
import time
from os import environ

from openpyxl import Workbook
import openpyxl

from PyQt5 import QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *

###-----CONVERTED PAGES---------###

from branch1 import *
from branch2 import *
from branch3 import *
from branch4 import *

def suppress_qt_warnings():
    environ["QT_DEVICE_PIXEL_RATIO"] = "0"
    environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    environ["QT_SCREEN_SCALE_FACTORS"] = "1"
    environ["QT_SCALE_FACTOR"] = "1"

suppress_qt_warnings()

Uygulama = QApplication(sys.argv)

###----------OPENING---------###
WinOpen = QMainWindow()
ui = Ui_MainWindow()
ui.setupUi(WinOpen)
###----------MAIN---------###
Win2= QMainWindow()
ui1=Ui_MainWindow1()
ui1.setupUi(Win2)
###----------SOLAR GLOSSARY---------###
Win3=QWidget()
ui3=Ui_Dialog()
ui3.setupUi(Win3)
###----------WIND GLOSSARY---------###
Win4=QWidget()
ui4=Ui_Dialog1()
ui4.setupUi(Win4)

WinOpen.show()

dest="#System Directory that contains program files#"

ui.lbl_map.setPixmap(QPixmap(dest+"MapTR.jpg"))

file = dest+"Solar_Report1.xlsx"
wb = openpyxl.load_workbook(file, read_only=True)
ws = wb.active

for i in ws.iter_rows(min_col=2,max_col=2,min_row=2):
    for j in i:
        #print(ws.cell(row=j.row, column=2).value)
        ui.cmbx_select.addItem(ws.cell(row=j.row, column=2).value)

province=[]

def cmbx_select():
    text = ui.cmbx_select.currentText()
    add(text)

def add(AA):
    province.clear()
    province.append(AA)
    ui.lnedt_selected.setText("Selected Province is {}".format(province[0]))

def wipe():
    ui.lnedt_selected.clear()

def menubar():
    ui1.menu_gloss_solar.triggered.connect(lambda: Win3.show())
    ui3.pushButton.clicked.connect(lambda: Win3.close())

    ui1.menu_gloss_wind.triggered.connect(lambda: Win4.show())
    ui4.pushButton.clicked.connect(lambda: Win4.close())

def map_butons():

###----------------------FUNCTION CONNECTIONS------------------###

    ui.pshbtn_run.clicked.connect(lambda: cmbx_select())

    ui.pshbtn_run.clicked.connect(lambda: result_solar(province[0]))
    ui.pshbtn_run.clicked.connect(lambda: result_wind(province[0]))

    ui.pshbtn_run.clicked.connect(lambda: hybrid_datas())

    ui.pshbtn_run.clicked.connect(lambda: WinOpen.close())
    ui.pshbtn_run.clicked.connect(lambda: Win2.show())

    ui1.pshbtn_back_to_main.clicked.connect(lambda: Win2.close())
    ui1.pshbtn_back_to_main.clicked.connect(lambda: WinOpen.show())

    ui1.pshbtn_back_to_main.clicked.connect(lambda: wipe())

    ui1.pshbtn_cikis.clicked.connect(lambda: sys.exit(Uygulama.exec_()))

    ui.pshbtn_runmap.clicked.connect(lambda: result_solar(province[0]))
    ui.pshbtn_runmap.clicked.connect(lambda: result_wind(province[0]))
    ui.pshbtn_runmap.clicked.connect(lambda: hybrid_datas())

    ui.pshbtn_runmap.clicked.connect(lambda: WinOpen.close())
    ui.pshbtn_runmap.clicked.connect(lambda: Win2.show())

###-----------------------MAP BUTTONS------------------------###

    ui.map_artvin.clicked.connect(lambda: add("Artvin"))
    ui.map_artvin.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_ankara.clicked.connect(lambda: add("Ankara"))
    ui.map_ankara.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_antalya.clicked.connect(lambda: add("Antalya"))
    ui.map_antalya.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_avcilar.clicked.connect(lambda:add("Avcılar"))
    ui.map_avcilar.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_bingol.clicked.connect(lambda: add("Bingöl"))
    ui.map_bingol.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_bodrum.clicked.connect(lambda: add("Bodrum"))
    ui.map_bodrum.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_buca.clicked.connect(lambda: add("Buca"))
    ui.map_buca.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_bursa.clicked.connect(lambda: add("Bursa"))
    ui.map_bursa.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_caanakkale.clicked.connect(lambda: add("Çanakkale"))
    ui.map_caanakkale.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_diyarbakir.clicked.connect(lambda:add("Diyarbakır"))
    ui.map_diyarbakir.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_edirne.clicked.connect(lambda: add("Edirne"))
    ui.map_edirne.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_erzincan.clicked.connect(lambda: add("Erzincan"))
    ui.map_erzincan.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_erzurum.clicked.connect(lambda: add("Erzurum"))
    ui.map_erzurum.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_eskisehir.clicked.connect(lambda: add("Eskişehir"))
    ui.map_eskisehir.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_gaziantep.clicked.connect(lambda: add("Gaziantep"))
    ui.map_gaziantep.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_hakkari.clicked.connect(lambda: add("Hakkari"))
    ui.map_hakkari.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_hatay.clicked.connect(lambda: add("Hatay"))
    ui.map_hatay.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_kadikoy.clicked.connect(lambda: add("Kadıköy"))
    ui.map_kadikoy.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_kars.clicked.connect(lambda: add("Kars"))
    ui.map_kars.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_kayseri.clicked.connect(lambda: add("Kayseri"))
    ui.map_kayseri.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_konya.clicked.connect(lambda: add("Konya"))
    ui.map_konya.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_malatya.clicked.connect(lambda: add("Malatya"))
    ui.map_malatya.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_mersin.clicked.connect(lambda: add("Mersin"))
    ui.map_mersin.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_ordu.clicked.connect(lambda: add("Ordu"))
    ui.map_ordu.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_rize.clicked.connect(lambda: add("Rize"))
    ui.map_rize.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_sakarya.clicked.connect(lambda: add("Sakarya"))
    ui.map_sakarya.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_sinop.clicked.connect(lambda: add("Sinop"))
    ui.map_sinop.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_sivas.clicked.connect(lambda: add("Sivas"))
    ui.map_sivas.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_tekirdag.clicked.connect(lambda: add("Tekirdağ"))
    ui.map_tekirdag.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_trabzon.clicked.connect(lambda: add("Trabzon"))
    ui.map_trabzon.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_usak.clicked.connect(lambda: add("Uşak"))
    ui.map_usak.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_van.clicked.connect(lambda: add("Van"))
    ui.map_van.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_zonguldak.clicked.connect(lambda: add("Zonguldak"))
    ui.map_zonguldak.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_corum.clicked.connect(lambda: add("Çorum"))
    ui.map_corum.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_adana.clicked.connect(lambda: add("Adana"))
    ui.map_adana.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

    ui.map_sanlurfa.clicked.connect(lambda: add("Şanlıurfa"))
    ui.map_sanlurfa.clicked.connect(lambda: print("Seçilen Şehir: {}".format(province[0])))

map_butons()
menubar()

def result_solar(datas):
    resultsolarlist = []

    file =dest+"Solar_Report1.xlsx"
    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active

    for row in ws.iter_rows(1):
        for cell in row:
            if cell.value == datas:
                resultsolarlist.append(ws.cell(row=cell.row, column=2).value)  #Province
                resultsolarlist.append(ws.cell(row=cell.row, column=3).value)  #GeoCoord
                resultsolarlist.append(ws.cell(row=cell.row, column=4).value)  #PVOUT
                resultsolarlist.append(ws.cell(row=cell.row, column=5).value)  #DNI
                resultsolarlist.append(ws.cell(row=cell.row, column=6).value)  #GHI
                resultsolarlist.append(ws.cell(row=cell.row, column=7).value)  #DHI
                resultsolarlist.append(ws.cell(row=cell.row, column=8).value)  ##GTIOpta
                resultsolarlist.append(ws.cell(row=cell.row, column=9).value)  #TEMP
                resultsolarlist.append(ws.cell(row=cell.row, column=10).value) #OPTA
                resultsolarlist.append(ws.cell(row=cell.row, column=11).value) #ELE

    solar_datas(resultsolarlist)

    return resultsolarlist

def solar_datas(list):
    obj=list

    ui1.lnedt_province.setText(obj[0])  #Province
    ui1.lnedt_coord.setText(obj[1])  #GeoCoord

    ui1.lnedt_pvout.setText(str(obj[2]))  #PVOUT
    ui1.lnedt_dni.setText(str(obj[3]))   #DNI
    ui1.lnedt_ghi.setText(str(obj[4]))   #GHI
    ui1.lnedt_dhi.setText(str(obj[5]))   #DHI
    ui1.lnedt_gtiopta.setText(str(obj[6]))  #GTIOpta
    ui1.lnedt_temp.setText(str(obj[7]))   #TEMP
    ui1.lnedt_opta.setText(str(obj[8]))   #OPTA
    ui1.lnedt_ele.setText(str(obj[9]))   #ELE

    province_pic=str(obj[0] + ".png")
    ui1.label_5.setPixmap(QPixmap(dest+province_pic))

def result_wind(datas):
    sonucwindlist = []

    file = dest+"Wind_Report1.xlsx"
    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active

    for row in ws.iter_rows(1):
        for cell in row:
            if cell.value == datas:
                sonucwindlist.append(ws.cell(row=cell.row, column=2).value)  # POUT

                sonucwindlist.append(ws.cell(row=cell.row, column=4).value)  # PVOUT
                sonucwindlist.append(ws.cell(row=cell.row, column=5).value)  # WS
                sonucwindlist.append(ws.cell(row=cell.row, column=6).value)  # h

    wind_datas(sonucwindlist)
    return sonucwindlist

def wind_datas(list):
    objR=list
    ui1.lnedt_pout.setText(str(objR[1]))
    ui1.lnedt_ws.setText(str(objR[2]))
    ui1.lnedt_h.setText(str(objR[3]))

    province_pic = str(objR[0] + "R.png")
    ui1.label_6.setPixmap(QPixmap(dest+province_pic))

def hybrid_datas():
    objS=result_solar(province[0])
    objR=result_wind(province[0])

    deger= str(objS[2])+" kWh/kWp "+"+ "+str(objR[1])+" W/m² "

    ui1.lnedt_pvout_top.setText(str(deger))
    ui1.lnedt_dni_hyb.setText(str(objS[3]))
    ui1.lnedt_ghi_hyb.setText(str(objS[4]))
    ui1.lnedt_ws_hyb.setText(str(objR[2]))
    ui1.lnedt_pvout_hyb.setText(str(objS[2]))
    ui1.lnedt_pout_hyb.setText(str(objR[1]))

sys.exit(Uygulama.exec_())



